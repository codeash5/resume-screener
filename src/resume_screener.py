import os
import re
import PyPDF2
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except:
    pass

# ---------- CONFIG ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def find_folder(folder_name):
    path_1 = os.path.join(BASE_DIR, "data", folder_name)
    if os.path.exists(path_1):
        return path_1
    path_2 = os.path.join(os.path.dirname(BASE_DIR), "data", folder_name)
    if os.path.exists(path_2):
        return path_2
    return None

RESUMES_FOLDER = find_folder("resumes")
JOBS_FOLDER = find_folder("job_descriptions")
TOP_N = 5  # Top matches per job

# ---------- Utility functions ----------
def extract_text_from_pdf(pdf_path):
    text = ""
    try:
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() or ""
    except Exception as e:
        print(f" Error reading {pdf_path}: {e}")
    return text

def clean_text(text):
    text = text.lower()
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'[^a-z0-9\s]', '', text)
    return text

# ---------- Skill Extraction ----------
SKILLS = [

# Programming
"python","java","c","c++","javascript",

# Data
"sql","mysql","postgresql",
"excel","power bi","tableau",

# Data science
"pandas","numpy","scikit-learn",
"machine learning","deep learning",

# Software engineering
"data structures","algorithms","git",
"docker","rest","api","microservices",

# Marketing
"digital marketing","seo","google analytics",
"social media","content writing",
"email marketing","campaign management"

]

def extract_skills(text, skill_list):

    text = text.lower()
    found_skills = []

    for skill in skill_list:

        pattern = r'\b' + re.escape(skill) + r'\b'

        if re.search(pattern, text):
            found_skills.append(skill)

    return found_skills

# ---------- Extract skills from Job Description ----------
def extract_jd_skills(jd_text):

    jd_skills = []

    for skill in SKILLS:
        pattern = r'\b' + re.escape(skill) + r'\b'
        if re.search(pattern, jd_text):
            jd_skills.append(skill)

    return jd_skills

def compute_skill_match(resume_skills, jd_skills):

    if not jd_skills:
        return 0

    matched = set(resume_skills) & set(jd_skills)

    return len(matched) / len(jd_skills)


# ---------- Load resumes (recursive) ----------
def load_resumes():
    resumes = {}
    if not RESUMES_FOLDER:
        print(" Could not find 'resumes' folder")
        return resumes

    print(f" Searching for PDFs in {RESUMES_FOLDER} (recursive)...")
    for root, _, files in os.walk(RESUMES_FOLDER):
        for file in files:
            if file.lower().endswith(".pdf"):
                path = os.path.join(root, file)
                relative_path = os.path.relpath(path, RESUMES_FOLDER)
                text = clean_text(extract_text_from_pdf(path))
                category = relative_path.split(os.sep)[0]
                if text.strip():
                    resumes[relative_path] = (text, category)
                else:
                    print(f"Skipping {relative_path} (no text extracted)")

    if resumes:
        print(f" Loaded {len(resumes)} resumes successfully.")

    else:
        print(f" No valid resumes found in {RESUMES_FOLDER}")
    return resumes

# ---------- Load job descriptions ----------
def load_job_descriptions():
    jobs = {}
    if not JOBS_FOLDER:
        print(" Could not find 'job_descriptions' folder")
        return jobs

    files = [f for f in os.listdir(JOBS_FOLDER) if f.lower().endswith(".txt")]
    for file in files:
        path = os.path.join(JOBS_FOLDER, file)
        with open(path, 'r', encoding='utf-8') as f:
            jobs[file] = clean_text(f.read())

    if jobs:
        print(f" Loaded {len(jobs)} job descriptions:")
        for j in jobs.keys():
            print(f"   - {j}")
    else:
        print(f" No job descriptions found in {JOBS_FOLDER}")
    return jobs

# ---------- Matching function ----------
def match_resumes_to_jobs(resumes, jobs):
    all_results = []
    for job_name, job_text in jobs.items():
        jd_skills = extract_jd_skills(job_text)
        print("Detected JD skills:", jd_skills)
        print(f"\n Job: {job_name}")
        resume_texts = [v[0] for v in resumes.values()]
        all_texts = [job_text] + resume_texts
        vectorizer = TfidfVectorizer(
        stop_words='english',
        ngram_range=(1,2),
        max_features=5000)
        tfidf_matrix = vectorizer.fit_transform(all_texts)
        similarities = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:]).flatten()
        ranked = []

        for name, score in zip(resumes.keys(), similarities):

            resume_text = resumes[name][0]
            category = resumes[name][1]

            resume_skills = extract_skills(resume_text, SKILLS)

            skill_score = compute_skill_match(resume_skills, jd_skills)
            missing_skills = [skill for skill in jd_skills if skill not in resume_skills][:5]

            final_score = (0.7 * score) + (0.3 * skill_score)

            ranked.append((name, category, score, skill_score, final_score, missing_skills))

        ranked = sorted(ranked, key=lambda x: x[4], reverse=True)
        for resume_name, category, score, skill_score, final_score, missing_skills in ranked[:TOP_N]:
            print(f"   {resume_name} ({category}) → sim:{score:.2f} skill:{skill_score:.2f} final:{final_score:.2f}")
        for rank, (resume_name, category, score, skill_score, final_score, missing_skills) in enumerate(ranked, start=1):
            all_results.append([
            job_name,
            resume_name,
            category,
            round(score,4),
            round(skill_score,4),
            round(final_score,4),
            round(final_score*100,2),
            rank,
            ", ".join(missing_skills)
        ])
    return all_results

# ---------- Save Excel ----------
def save_to_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resume Scores"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for jd in df["Job Description"].unique():
        jd_rows = [i+2 for i, val in enumerate(df["Job Description"]) if val == jd]
        jd_ranks = df[df["Job Description"]==jd]["Rank"].tolist()
        for i, rank in zip(jd_rows, jd_ranks):
            cell = ws.cell(row=i, column=8)  # Rank column
            if rank <=2:
                cell.fill = green
            elif rank <=5:
                cell.fill = yellow
            else:
                cell.fill = red

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_len

    wb.save(path)
    print(f" Excel saved to: {path}")

# ---------- Main ----------
if __name__ == "__main__":
    print(f" Current working directory: {os.getcwd()}")
    print(f" Resumes folder path: {RESUMES_FOLDER or 'Not found'}")
    print(f" Jobs folder path: {JOBS_FOLDER or 'Not found'}")

    resumes = load_resumes()
    jobs = load_job_descriptions()

    if not resumes or not jobs:
        print(" Cannot proceed - missing resumes or job descriptions.")
    else:
        results = match_resumes_to_jobs(resumes, jobs)
        df = pd.DataFrame(results, columns=[
        "Job Description",
        "Resume Filename",
        "Category",
        "Similarity Score",
        "Skill Match",
        "Final Score",
        "Match %",
        "Rank",
        "Missing Skills"
        ])
        csv_path = os.path.join(os.path.dirname(BASE_DIR), "data", "resume_scores_per_jd.csv")
        df.to_csv(csv_path, index=False)
        print(f" CSV saved to: {csv_path}")
        excel_path = os.path.join(os.path.dirname(BASE_DIR), "data", "resume_scores_per_jd.xlsx")
        save_to_excel(df, excel_path)




